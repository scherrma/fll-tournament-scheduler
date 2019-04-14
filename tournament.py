#!/usr/bin/env python3
"""A module containing the Tournament class for using in creating FLL qualifier schedules."""
from datetime import datetime, timedelta
import math
import os
import sys
import tkinter
from tkinter import filedialog
import pandas
import openpyxl
import openpyxl.styles as styles
import scheduler.util as util
from scheduler.team import Team
import scheduler.min_cost

class Tournament:
    """A class designed to create and export schedules for FLL qualifier tournaments."""
    def __init__(self):
        """Creates a tournament and requests a roster/settings file if one was not provided."""
        if len(sys.argv) == 1:
            root = tkinter.Tk()
            root.withdraw()
            self.fpath = filedialog.askopenfilename(
                filetypes=[("Excel files (*.xls, *.xlsm, *.xlsx)", "*.xls *.xlsm *.xlsx")])
            root.destroy()
        elif len(sys.argv) == 2:
            self.fpath = sys.argv[1]
        else:
            print("{} accepts 0 or 1 arguments; {} received".format(sys.argv[0], len(sys.argv) - 1))

    def schedule(self):
        """Top-level scheduling function; reads data, generates schedule, and exports."""
        try:
            print("Reading data")
            self.read_data(self.fpath)
            for team in self.teams:
                team.add_event(*self.coach_meet, 0, 0)
                team.add_event(*self.opening, 1, 0)

            print("Starting judge scheduling")
            self.split_divisions()
            if self.scheduling_method == "Interlaced":
                self.schedule_interlaced()
            elif self.scheduling_method == "Block":
                self.schedule_block()
            else:
                raise ValueError(self.scheduling_method + " scheduling is not supported")
            print("Scheduling competition tables")
            self.assign_tables()
            print("Exporting schedule")
            self.export()

        except Exception as excep:
            raise excep
            #print(excep)

        if sys.platform == "win32":
            os.system("pause")

    def schedule_interlaced(self):
        """Top-level function controlling judge and table schedules for interlaced tournaments."""
        self.judge_interlaced()

        time_start = [e_start + e_length + self.travel for (e_start, e_length, *_)
                      in self.teams[3*self.j_calib].events]
        time_start = sorted([t for t in time_start if t >= self.j_start])
        time_start = next((time for time in time_start if self.teams[3*self.j_calib]
                           .available(time, self.t_duration[0], self.travel)))
        team_idx = next((t for t in range(3*self.j_calib + 1) if
                         self._team(t).available(time_start, self.t_duration[0], self.travel)))

        run_rate = 3*self.j_sets*self.t_duration[0]/(self.j_duration + self.j_break/self.j_consec)

        min_early_run_rate = max(2, 2*math.ceil(run_rate/2))
        if self.num_teams % min_early_run_rate:
            early_avail = [self._team(team_idx - i).available(time_start - self.t_duration[0],
                                                              self.t_duration[0], self.travel)
                           for i in range(min_early_run_rate)]
            if all(early_avail):
                team_idx -= len(early_avail)
                time_start -= self.t_duration[0]

        self.t_slots = []
        t_rounds = len(self.event_names) - 5
        time_start = self.schedule_matches(time_start, team_idx, (0, 1)[:t_rounds], run_rate)
        if t_rounds > 1:
            self.t_slots += [None]
            time_restart = [sum(self._team(t).events[-1][:2], self.travel
                                - t // (2*self.t_pairs) * self.t_duration[2])
                            for t in range(self.num_teams)]
            time_start = max(time_restart + [time_start + self.t_lunch_duration])
            self.schedule_matches(time_start, 0, range(2, t_rounds), None)

    def judge_interlaced(self):
        """Generates the judging schedule for tournaments using interlaced scheduling."""
        self.j_calib = self.j_calib and not self.divisions

        max_room = max([math.ceil(len(teams) / rooms) for teams, rooms in self.divs])
        most_rooms = max([rooms for teams, rooms in self.divs])

        picks = [util.rpad(rooms*[i], most_rooms, None) for i, (tms, rooms) in enumerate(self.divs)]
        pick_order = math.ceil(max_room / 3)*[val for div_picks in zip(*picks)
                                              for val in 3*div_picks if val is not None]

        for div, (teams, rooms) in enumerate(self.divs):
            last_team = util.nth_occurence(pick_order, div, len(teams))
            pick_order = [x for idx, x in enumerate(pick_order) if idx <= last_team or x != div]

        excess = min([-len(teams) % 3 for teams, rooms in self.divs if
                      math.ceil(len(teams) / rooms) == max_room])
        div_teams = [util.rpad([i for i, j in enumerate(pick_order) if j == div],
                               3*math.ceil(max_room/3)*rooms - excess, None)
                     for div, (teams, rooms) in enumerate(self.divs)]

        team_order = [list(zip(idxs, teams)) for idxs, (teams, rooms) in zip(div_teams, self.divs)]
        self.teams = [team for idx, team in sorted(sum(list(team_order), []))]

        self.j_slots = [[], [], []]
        for teams, (_, rooms) in zip(div_teams, self.divs):
            rot_dir = 1 if (3*self.j_calib - len(teams)) % 3 == 1 else -1
            tmp = [sum([teams[(j + i*rot_dir) % 3 : len(teams) : 3] for i in range(3)],
                       [])[self.j_calib:] for j in range(3)]
            for i in range(3):
                self.j_slots[i] += [tmp[i][j::rooms] for j in range(rooms)]
        self.j_slots = [[util.rpad(room, max_room, None) for room in cat] for cat in self.j_slots]
        self.j_slots = list(zip(*[list(zip(*cat)) for cat in self.j_slots]))
        if self.j_calib:
            self.j_slots = [([0], [1], [2])] + self.j_slots
        self.assign_judge_times()

    def schedule_block(self):
        """Generates judging and table schedules using block scheduling."""
        raise NotImplementedError("Block scheduling is not implemented yet")

    def split_divisions(self):
        """Sets self.divs to a list of (teams, rooms for those teams) based on division."""
        if not self.divisions:
            self.divs = [(self.teams, self.j_sets)]
        room_max = max(12, math.ceil(self.num_teams / self.j_sets))

        self.divs = [(teams, math.ceil(len(teams) / room_max)) for teams in
                     [[team for team in self.teams if team.div == div] for div in
                      {team.div for team in self.teams}]]
        self.divs.sort(key=lambda div: -len(div[0]) % room_max)

        if sum([rooms for _, rooms in self.divs]) > self.j_sets:
            room_divs, impure_divs = [], []
            teams_left, rooms_left = self.num_teams, self.j_sets
            for teams, rooms in self.divs:
                if (-len(teams) // room_max + rooms_left) * room_max >= teams_left - len(teams):
                    room_divs.append((teams, rooms))
                    teams_left -= len(teams)
                    rooms_left -= rooms
                else:
                    impure_divs += [teams]

            room_max = math.ceil(teams_left / rooms_left)
            excess = rooms_left*room_max - teams_left

            idx, spillover = 0, 0
            impure_teams = sum(impure_divs, [])
            for i, teams in enumerate(impure_divs):
                skips = int(i < excess % len(impure_divs)) + (excess // len(impure_divs))
                pure = ((len(teams) - spillover) // room_max) * room_max - max(0, skips - 1)
                room_divs.append((impure_teams[idx:idx + pure], math.ceil(pure / room_max)))

                idx += pure
                room_divs.append((impure_teams[idx:idx + room_max - bool(skips)], 1))
                idx += room_max - bool(skips)
                spillover += pure + room_max - bool(skips) - len(teams)
            self.divs = [(teams, rooms) for teams, rooms in room_divs if teams]

    def assign_judge_times(self):
        """Determines when each judging session will happen and assigns teams to those slots."""
        breaks = {0, len(self.j_slots)} | set(range(self.j_calib, len(self.j_slots), self.j_consec))
        breaks = sorted(list(breaks))
        times = [[self.j_start + bool(j and self.j_calib)*self.travel
                  + max(i - self.j_calib, 0)*self.j_break + j*self.j_duration
                  for j in range(breaks[i], breaks[i+1] + 1)] for i in range(len(breaks) - 1)]
        j_blockers = [(self.j_lunch, self.j_lunch_duration)]\
                   + [(start - self.travel, duration + 2*self.travel) for start, duration in
                      (self.opening, self.coach_meet)]
        for start, length in sorted(j_blockers):
            delay = max(timedelta(0), min(length, start + length - times[0][0]))
            delay -= self.j_break if start >= times[0][-1] else timedelta(0)
            times = [[time + (start < cycle[-1])*delay for time in cycle] for cycle in times]
        times = [time for cycle in times for time in cycle]

        for breaktime in breaks[-2:0:-1]:
            self.j_slots.insert(breaktime, None)
        self.j_slots = list(zip(times, self.j_slots))
        for time, teams in filter(lambda x: x[1] is not None, self.j_slots):
            for cat, cat_teams in enumerate(teams):
                for room, team in filter(lambda x: x[1] is not None, enumerate(cat_teams)):
                    self.teams[team].add_event(time, self.j_duration_team, cat + 2, room)

    def schedule_matches(self, time, team, rounds, run_rate):
        """Determines when table matches will occur and assigns teams to matches."""
        def avail(start, rnd):
            return [(t, self._team(start + t).available(time, self.t_duration[rnd], self.travel))
                    for t in range(self.num_teams)]
        start_team = team
        if run_rate is None:
            match_sizes = [2*self.t_pairs]
        else:
            run_rate = 2*min(math.ceil(run_rate/2), self.t_pairs)
            match_sizes = [max(2, run_rate - 2), run_rate]

        while team < len(rounds)*self.num_teams + start_team:
            rnd = rounds[(team - start_team) // self.num_teams]
            max_teams = next((t for t, free in avail(team, rnd) if not free),
                             len(rounds)*self.num_teams)
            max_matches = (self._team(team).next_event(time)[0] - time - self.travel)\
                    // self.t_duration[rnd]

            if team + max_teams >= len(rounds)*self.num_teams + start_team:
                max_teams = len(rounds)*self.num_teams + start_team - team
                max_matches = min(max_matches, math.ceil(max_teams / match_sizes[-1]))
            else:
                max_matches = min(math.ceil((len(rounds)*self.num_teams + start_team - team)
                                            / match_sizes[-1]), max_matches)
                max_teams -= max_teams % 2

            if max_matches == 0:
                time += self.t_duration[rnd]

            next_matches = util.sum_to(match_sizes, max_teams, max_matches, force_take_all=
                                       team + max_teams - start_team >= len(rounds)*self.num_teams)
            while sum(next_matches[:-1]) + (team - start_team) % self.num_teams > self.num_teams:
                next_matches.pop()

            for match_size in next_matches:
                timeslot = [t % self.num_teams for t in range(team, team + match_size)]
                self.t_slots += [(time, rnd, util.rpad(timeslot, match_sizes[-1], None))]
                team += match_size
                time += self.t_duration[rnd]
        return time

    def assign_tables(self, assignment_passes=2):
        """Reorders the teams in self.t_slots to minimize table repetition for teams."""
        prev_tables = [[0 for i in range(2*self.t_pairs)] for j in range(self.num_teams)]
        def cost(order):
            return sum(prev_tables[team][table]**1.1 for table, team in enumerate(order)
                       if team is not None)

        for assign_pass in range(assignment_passes):
            for(time, rnd, teams) in filter(None, self.t_slots):
                if assign_pass:
                    for table, team in filter(lambda x: x[1] is not None, enumerate(teams)):
                        prev_tables[team][table] -= 1
                teams[:] = scheduler.min_cost.min_cost(teams, cost)
                for table, team in filter(lambda x: x[1] is not None, enumerate(teams)):
                    prev_tables[team][table] += 1
                    if assign_pass + 1 == assignment_passes:
                        teams[:] = util.rpad(teams, 2*self.t_pairs, None)
                        team_rnd = sum(1 for event in self._team(team).events if event[2] > 4)
                        self._team(team).add_event(time, self.t_duration[rnd], 5 + team_rnd, table)

    def export(self):
        """Exports schedule to an xlsx file; uses the tournament name for the file name."""
        time_fmt = ('%#I:%M %p' if sys.platform == "win32" else '%-I:%M %p')
        team_width = len(self.teams[0].info(self.divisions))
        workbook = openpyxl.load_workbook(self.fpath)

        for sheet in [ws for ws in workbook.sheetnames if ws != 'Team Information']:
            del workbook[sheet]

        self._export_judge_views(workbook, time_fmt, team_width)
        self._export_table_views(workbook, time_fmt, team_width)
        self._export_team_views(workbook, time_fmt)

        outfpath = os.path.join(os.path.dirname(self.fpath),
                                self.tournament_name.lower().replace(' ', '_') + '_schedule')
        saved, count = False, 0
        while not saved:
            try:
                workbook.save('{}{}.xlsx'.format(outfpath, ' ({})'.format(count) if count else ''))
                saved = True
            except PermissionError:
                count += 1
        print('Schedule saved: {}{}.xlsx'.format(outfpath, ' ({})'.format(count) if count else ''))

    def _export_judge_views(self, workbook, time_fmt, team_width):
        """Adds the four judging-focused sheets to the output workbook."""
        thin = styles.Side(border_style='thin', color='000000')
        thick = styles.Side(border_style='thick', color='000000')

        ws_overall = workbook.create_sheet("Judging Rooms")
        cat_sheets = [workbook.create_sheet(cat) for cat in self.event_names[2:5]]

        header_total, rooms_total = [], []
        for i in range(3):
            header_part = [self.event_names[i + 2]] + (self.j_sets*team_width - 1)*['']
            rooms_part = sum([[room] + (team_width - 1)*[''] for room in self.rooms[i + 2]], [])

            header_total += header_part
            rooms_total += rooms_part

            cat_sheets[i].append([''] + header_part)
            cat_sheets[i].append([''] + rooms_part)
        ws_overall.append([''] + header_total)
        ws_overall.append([''] + rooms_total)

        for time, teams in self.j_slots:
            line = [time.strftime(time_fmt)]
            if teams is not None:
                teams = [[None if t is None else self.teams[t] for t in cat] for cat in teams]
                if len(teams[0]) == 1 and self.j_calib:
                    line += sum([teams[i][0].info(self.divisions)
                                 + ["all {} judges in {}".format(self.event_names[i + 2].lower(),
                                                                 self.rooms[i + 2][0])]
                                 + (team_width*(self.j_sets - 1) - 1)*[''] for i in range(3)], [])
                else:
                    line += sum([['']*(team_width - 1) + ['None'] if team is None else
                                 team.info(self.divisions) for cat in teams for team in cat], [])
            ws_overall.append(line)
            for i in range(3):
                cat_sheets[i].append([line[0]] + line[i*team_width*self.j_sets + 1:
                                                      (i + 1)*team_width*self.j_sets + 1])

        col_sizes = [1 + max(len(str(text)) for text in cat) for cat in 
                     zip(*[team.info(self.divisions) for team in self.teams])]
        for sheet in [ws_overall] + cat_sheets: #formatting
            util.basic_ws_format(sheet, 4)
            for col in sheet.columns:
                if col[0].column > 1:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width =\
                            col_sizes[(col[0].column - 2) % len(col_sizes)]
            util.ws_borders(sheet, ((styles.Border(left=thick), team_width*self.j_sets, 2, 0),
                                    (styles.Border(left=thin), team_width, 2, 1 + 2*self.j_calib)))

            for i in range((len(list(sheet.columns)) - 1) // (team_width*self.j_sets)):
                sheet.cell(row=1, column=team_width*self.j_sets*i + 2).font = styles.Font(bold=True)
                sheet.merge_cells(start_row=1, start_column=2 + i*team_width*self.j_sets,
                                  end_row=1, end_column=1 + (i + 1)*team_width*self.j_sets)
                for j in range(self.j_sets):
                    sheet.merge_cells(start_row=2, start_column=team_width*(self.j_sets*i + j) + 2,
                                      end_row=2, end_column=team_width*(self.j_sets*i + j + 1) + 1)
                if self.j_calib:
                    sheet.merge_cells(start_row=3, start_column=2 + team_width*(self.j_sets*i + 1),
                                      end_row=3, end_column=1 + team_width*(self.j_sets*(i + 1)))

    def _export_table_views(self, workbook, time_fmt, team_width):
        """Adds the competition table focused sheets to the output workbook."""
        thin = styles.Side(border_style='thin', color='000000')
        thick = styles.Side(border_style='thick', color='000000')

        sheet_overall = workbook.create_sheet("Competition Tables")
        header = sum([[tbl] + (team_width - 1)*[''] for tbl in self.rooms[5]], [''])
        sheet_overall.append(header)
        t_pair_sheets = [workbook.create_sheet(room[:-2]) for room in self.rooms[5][::2]]
        for t_pair in range(self.t_pairs):
            t_pair_sheets[t_pair].append([''] + header[2*team_width*t_pair + 1:
                                                       2*team_width*(t_pair + 1)])

        for slot in self.t_slots:
            if slot is None:
                for sheet in t_pair_sheets + [sheet_overall]:
                    sheet.append([''])
            elif all([team is None for team in slot[2]]):
                for sheet in t_pair_sheets + [sheet_overall]:
                    sheet.append([slot[0].strftime(time_fmt)])
            else:
                line = sum([(team_width - 1)*[''] + ['None'] if t is None else
                            self.teams[t].info(self.divisions) for t in slot[2]],
                           [slot[0].strftime(time_fmt)])
                sheet_overall.append(line)
                for t_pair in range(self.t_pairs):
                    t_pair_sheets[t_pair].append([line[0]] + line[2*team_width*t_pair + 1:
                                                                  2*team_width*(t_pair + 1) + 1])

        col_sizes = [1 + max(len(str(text)) for text in cat) for cat in 
                     zip(*[team.info(self.divisions) for team in self.teams])]
        for sheet in [sheet_overall] + t_pair_sheets:
            util.basic_ws_format(sheet, 2)
            for col in sheet.columns:
                if col[0].column > 1:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width =\
                            col_sizes[(col[0].column - 2) % len(col_sizes)]
            util.ws_borders(sheet, ((styles.Border(left=thick), 2*team_width, 2, 0),
                                    (styles.Border(left=thin), 2*team_width, 2 + team_width, 0)))
            for i in range(2, len(list(sheet.columns)), team_width):
                sheet.merge_cells(start_row=1, start_column=i,
                                  end_row=1, end_column=i + team_width - 1)

    def _export_team_views(self, workbook, time_fmt):
        """Adds event-sorted and time-sorted team-focused views to the output workbook."""
        ws_chron = workbook.create_sheet("Team View (Chronological)")
        ws_event = workbook.create_sheet("Team View (Event)")
        team_header = ['Team Number'] + (['Division'] if self.divisions else []) + ['Team Name']
        ws_chron.append(team_header + ['Event ' + str(i + 1) for i in range(len(self.event_names))])
        ws_event.append(team_header + self.event_names)

        for team in sorted(self.teams, key=lambda t: t.num):
            ws_chron.append(team.info(self.divisions)
                            + ['{} at {}, {}'.format(self.event_names[cat], time.strftime(time_fmt),
                                                     self.rooms[min(5, cat)][loc])
                               for (time, duration, cat, loc) in team.events])
            ws_event.append(team.info(self.divisions)
                            + ['{}, {}'.format(time.strftime(time_fmt),
                                               self.rooms[min(5, cat)][loc])
                               for (time, length, cat, loc)
                               in sorted(team.events, key=lambda x: x[2])])

        for sheet in (ws_chron, ws_event):
            util.basic_ws_format(sheet)

    def _team(self, team_num):
        """Returns the team at the specified internal index; wraps modularly."""
        return self.teams[team_num % self.num_teams]

    def read_data(self, fpath):
        """Imports the team roster and scheduling settings from the input form."""
        dfs = pandas.read_excel(fpath, sheet_name=["Team Information", "Input Form"])

        team_sheet = dfs["Team Information"]
        column_check = ["Team Number", "Team"]
        if all([x in team_sheet.columns for x in column_check]):
            self.divisions = ("Division" in team_sheet.columns)
            if self.divisions:
                column_check += ["Division"]
            self.teams = [Team(*x) for x in team_sheet.loc[:, column_check].values]
        else:
            raise KeyError("Could not find columns 'Team Number' and 'Team' in 'Team Information'")

        param_sheet = dfs["Input Form"]
        if any([x not in param_sheet.columns for x in ("key", "answer")]):
            raise KeyError("Could not find columns 'key' and 'answer' in sheet 'Input Form'")
        self.param_sheet = param_sheet.set_index("key")
        param = dict(self.param_sheet.loc[:, "answer"].items())

        self.num_teams = len(self.teams)
        try:
            self.tournament_name = param["tournament_name"]
            self.scheduling_method = param["scheduling_method"]
            self.travel = timedelta(minutes=param["travel_time"])
            self.event_names = ["Coaches' Meeting", "Opening Ceremonies", 'Project', 'Robot Design',
                                'Core Values']
            self.event_names += self.param_sheet.loc["t_round_names"].dropna().values.tolist()[1:]
            self.coach_meet = (datetime.combine(datetime(1, 1, 1), param["coach_start"]),
                               timedelta(minutes=param["coach_duration"]))
            self.opening = (datetime.combine(datetime(1, 1, 1), param["opening_start"]),
                            timedelta(minutes=param["coach_duration"]))

            self.j_start = datetime.combine(datetime(1, 1, 1), param["j_start"])
            self.j_sets = param["j_sets"]
            self.rooms = [[param["coach_room"]], [param["opening_room"]]]
            self.rooms += [self.param_sheet.loc[key].dropna().values.tolist()[1:]
                           for key in ("j_project_rooms", "j_robot_rooms", "j_values_rooms")]
            self.j_calib = (param["j_calib"] == "Yes")
            self.j_duration = timedelta(minutes=param["j_duration"])
            self.j_duration_team = timedelta(minutes=10)
            self.j_consec = param["j_consec"]
            self.j_break = timedelta(minutes=param["j_break"])
            self.j_lunch = datetime.combine(datetime(1, 1, 1), param["j_lunch"])
            self.j_lunch_duration = timedelta(minutes=param["j_lunch_duration"])

            self.t_pairs = param["t_pairs"]
            self.rooms += [sum([[tbl + ' A', tbl + ' B'] for tbl in
                                self.param_sheet.loc["t_pair_names"].dropna().values.tolist()[1:]],
                               [])][:2*self.t_pairs]
            self.t_duration = [timedelta(minutes=x) for x in
                               self.param_sheet.loc["t_durations"].dropna().values.tolist()[1:]]
            self.t_lunch = datetime.combine(datetime(1, 1, 1), param["t_lunch"])
            self.t_lunch_duration = timedelta(minutes=param["t_lunch_duration"])


        except KeyError as excep:
            raise KeyError(str(excep) + " not found in 'key' column in sheet 'Input Form'")

if __name__ == "__main__":
    print("Python environment initialized")
    Tournament().schedule()
