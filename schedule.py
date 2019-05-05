#!/usr/bin/env python3
"""Module for use in scheduling one-day FLL tournaments."""
from datetime import datetime, timedelta
import os
import sys
import math
import tkinter
from tkinter import filedialog
import pandas
from numpy import isnan
import openpyxl
import openpyxl.styles as styles
from openpyxl.utils import get_column_letter
import warnings
from scheduler.tournament import Tournament

def read_data(fpath):
    """Imports the team roster and scheduling settings from the input form."""
    dfs = pandas.read_excel(fpath, sheet_name=["Team Information", "Input Form"], dtype=object)

    team_sheet = dfs["Team Information"]
    column_check = ["Team Number", "Team"]
    if all([x in team_sheet.columns for x in column_check]):
        divisions = ("Division" in team_sheet.columns)
        if divisions:
            column_check += ["Division"]
        teams = team_sheet.loc[:, column_check].values
        team_info_base = [(i, list(team_sheet.columns).index(cat) + 1) for i, cat in
                          list(enumerate(column_check[::-1]))]
        #any time we print full team data just print the number and look the other info up
        team_info = ["=index(indirect(\"'Team Information'!C{1}\", false),"\
                      "match(indirect(\"RC[-{2}]\", false), "\
                      "indirect(\"'Team Information'!C{0}\", false), 0))"
                     .format(team_info_base[-1][1], cat, offset + 1) for offset, cat
                     in team_info_base[:-1]]
    else:
        raise KeyError("Could not find columns 'Team Number' and 'Team' in 'Team Information'")

    param_sheet = dfs["Input Form"]
    if any([x not in param_sheet.columns for x in ("key", "answer")]):
        raise KeyError("Could not find columns 'key' and 'answer' in sheet 'Input Form'")
    param_sheet = param_sheet.set_index("key")
    param = dict(param_sheet.loc[:, "answer"].items())

    try: #there are a lot of settings. this appears to be necessary
        tournament_name = param["tournament_name"]
        scheduling_method = param["scheduling_method"]
        travel = timedelta(minutes=param["travel_time"])
        event_names = ["Coaches' Meeting", "Opening Ceremonies", 'Project', 'Robot Design',
                       'Core Values']
        event_names += map(str, param_sheet.loc["t_round_names"].dropna().values.tolist()[1:])
        t_rounds = len(event_names) - 5
        coach_meet = (datetime.combine(datetime(1, 1, 1), param["coach_start"]),
                      timedelta(minutes=param["coach_duration"]))
        opening = (datetime.combine(datetime(1, 1, 1), param["opening_start"]),
                   timedelta(minutes=param["coach_duration"]))
        lunch = (datetime.combine(datetime(1, 1, 1), param["lunch_earliest"]),
                 datetime.combine(datetime(1, 1, 1), param["lunch_latest"]),
                 timedelta(minutes=param["lunch_duration"]))

        j_start = datetime.combine(datetime(1, 1, 1), param["j_start"])
        j_sets = param["j_sets"]
        rooms = [[param["coach_room"]], [param["opening_room"]]]
        rooms += [param_sheet.loc[key].dropna().values.tolist()[1:]
                  for key in ("j_project_rooms", "j_robot_rooms", "j_values_rooms")]
        j_calib = (param["j_calib"] == "Yes") and not divisions
        j_duration = (timedelta(minutes=param["j_duration"]), timedelta(minutes=10))
        j_breaks = (param["j_consec"] or len(teams), timedelta(minutes=param["j_break"]))

        t_pairs = param["t_pairs"]
        t_stagger = (param["t_stagger"] == 'Yes')
        t_consec = param["t_consec"] or len(teams)*t_rounds
        t_names = param_sheet.loc[["t_pair_names", "t_pair_names_second"]].iloc[:, 1:1 + t_pairs].T
        t_names = [[str(tbl) for tbl in row if not pandas.isnull(tbl)] for row in t_names.values]
        rooms += [sum([tbls if len(tbls) > 1 else [tbls[0] + ' A', tbls[0] + ' B']
                       for tbls in t_names], [])]
        t_duration = [timedelta(minutes=x) for x in
                      param_sheet.loc["t_durations"].dropna().values.tolist()[1:]]

    except KeyError as excep:
        raise KeyError(str(excep) + " not found in 'key' column in sheet 'Input Form'")

    return ((teams, divisions, scheduling_method, travel, coach_meet, opening, lunch, j_start,
             j_sets, j_calib, j_duration, j_breaks, t_rounds, t_pairs, t_stagger, t_consec,
             t_duration), tournament_name, (team_info, event_names, rooms, t_names))

def export(tment, workbook, team_info, event_names, rooms, tnames):
    """Exports schedule to an xlsx file; uses the tournament name for the file name."""
    print("Exporting schedule")
    for sheet in [ws for ws in workbook.sheetnames if ws != 'Team Information']:
        del workbook[sheet]

    time_fmt = "%{}I:%M %p".format('#' if sys.platform == "win32" else '-')
    export_judge_views(tment, workbook, time_fmt, team_info, event_names, rooms)
    export_table_views(tment, workbook, time_fmt, team_info, rooms, tnames)
    export_team_views(tment, workbook, time_fmt, team_info, event_names, rooms)

def export_judge_views(tment, workbook, time_fmt, team_info, event_names, rooms):
    """Adds the four judging-focused sheets to the output workbook."""
    thin = styles.Side(border_style='thin', color='000000')
    thick = styles.Side(border_style='thick', color='000000')
    team_width = 1 + len(team_info)

    sheets = [workbook.create_sheet(name) for name in ["Judging Rooms"] + event_names[2:5]]

    #writing data
    rows = [[['']], [['']]]
    for i in range(3):
        rows[0].append([event_names[i + 2]] + (tment.j_sets*team_width - 1)*[''])
        rows[1].append(sum([[room] + (team_width - 1)*[''] for room in rooms[i + 2]], []))

    for time, teams in tment.j_slots:
        rows.append([[time.strftime(time_fmt)]])
        if teams is not None:
            teams = [[None if t is None else tment.teams[t] for t in cat] for cat in teams]
            if len(teams[0]) == 1 and tment.j_calib:
                rows[-1] += [[teams[i][0].num] + team_info
                             + ["all {} judges in {}".format(event_names[i + 2].lower(),
                                                             rooms[i + 2][0])]
                             + (team_width*(tment.j_sets - 1) - 1)*[''] for i in range(3)]
            else:
                rows[-1] += [sum([['']*(team_width - 1) + ['None'] if team is None else
                                  [team.num] + team_info for team in cat], []) for cat in teams]
    for row in rows:
        sheets[0].append(sum(row, []))
        for i in range(3):
            sheets[i + 1].append(row[0] + (row[i + 1] if len(row) > 1 else []))

    #formatting - borders, cell merges, striped shading, etc
    col_sizes = [1 + max(len(str(text)) for text in cat) for cat in
                 zip(*[team.info(tment.divisions) for team in tment.teams])]
    for sheet in sheets:
        basic_sheet_format(sheet, 4)
        for col in list(sheet.columns)[1:]:
            sheet.column_dimensions[get_column_letter(col[0].column)].width =\
                    col_sizes[(col[0].column - 2) % len(col_sizes)]
        sheet_borders(sheet, ((styles.Border(left=thick), team_width*tment.j_sets, 2, 0),
                              (styles.Border(left=thin), team_width, 2, 1 + 2*tment.j_calib)))

        for i in range((len(list(sheet.columns)) - 1) // (team_width*tment.j_sets)):
            sheet.cell(row=1, column=team_width*tment.j_sets*i + 2).font = styles.Font(bold=True)
            sheet.merge_cells(start_row=1, start_column=2 + i*team_width*tment.j_sets,
                              end_row=1, end_column=1 + (i + 1)*team_width*tment.j_sets)
            for j in range(tment.j_sets):
                sheet.merge_cells(start_row=2, start_column=team_width*(tment.j_sets*i + j) + 2,
                                  end_row=2, end_column=team_width*(tment.j_sets*i + j + 1) + 1)
            if tment.j_calib:
                sheet.merge_cells(start_row=3, start_column=2 + team_width*(tment.j_sets*i + 1),
                                  end_row=3, end_column=1 + team_width*(tment.j_sets*(i + 1)))

def export_table_views(tment, workbook, time_fmt, team_info, rooms, tnames):
    """Adds the competition table focused sheets to the output workbook."""
    thin = styles.Side(border_style='thin', color='000000')
    thick = styles.Side(border_style='thick', color='000000')
    team_width = 1 + len(team_info)
    space = 2
    split = 1 + 2*team_width*((tment.t_pairs + 1) // 2)
    staggered = [tment.t_stagger and i > (tment.t_pairs - 1) // 2 for i in range(tment.t_pairs)]

    #writing data
    sheet_overall = workbook.create_sheet("Competition Tables")
    header = sum([[tbl] + (team_width - 1)*[''] for tbl in rooms[5]], [''])
    header[split:split] = tment.t_stagger*(space + 1)*['']
    sheet_overall.append(header)
    t_pair_sheets = [workbook.create_sheet('-'.join(tbls)) for tbls in tnames]
    for t_pair in range(tment.t_pairs):
        t_pair_sheets[t_pair]\
                .append([''] + header[2*team_width*t_pair + 1 + staggered[t_pair]*(space + 1):
                                      2*team_width*(t_pair + 1) + staggered[t_pair]*(space + 1)])

    for slot in tment.t_slots:
        if slot is None:
            for sheet in [sheet for sheet in t_pair_sheets + [sheet_overall] for i in range(2)]:
                sheet.append([''])
        elif all([team is None for team in slot[2]]):
            sheet_overall.append([slot[0][0].strftime(time_fmt)] + tment.t_stagger*((split + 1)*[''] +
                                 [slot[0][1].strftime(time_fmt)]))
            for i, sheet in enumerate(t_pair_sheets):
                sheet.append([slot[0][staggered[i]].strftime(time_fmt)])
        else:
            line = sum([(team_width - 1)*[''] + ['None'] if t is None else
                        [tment.teams[t].num] + team_info for t in slot[2]],
                       [slot[0][0].strftime(time_fmt)])
            line[split:split] = space*[''] + [slot[0][1].strftime(time_fmt)] if tment.t_stagger else []
            sheet_overall.append(line)
            for t_pair in range(tment.t_pairs):
                time_str = slot[0][staggered[t_pair]].strftime(time_fmt)
                ls_start = 2*team_width*t_pair + (space + 1)*staggered[t_pair] + 1
                t_pair_sheets[t_pair].append([time_str] + line[ls_start:ls_start + 2*team_width])

    #formatting - borders, cell merges, striped shading, etc
    col_wide = [-1, -1] + 2*tment.t_pairs*[1 + max(len(str(text)) for text in cat) for cat in
                                         zip(*[team.info(tment.divisions) for team in tment.teams])]
    col_wide[split + 1:split + 1] = tment.t_stagger*(space*[10] + [-1])
    for sheet in [sheet_overall] + t_pair_sheets:
        basic_sheet_format(sheet, 2)
        for col, width in [(col, width) for col, width in enumerate(col_wide) if width > 0]:
            sheet.column_dimensions[get_column_letter(col)].width = width
        thick_border = [1 + 2*i*team_width for i in range(tment.t_pairs + 1)]
        if tment.t_stagger:
            thick_border = [1 + 2*i*team_width for i in range(math.ceil(tment.t_pairs / 2) + 1)]
            thick_border += [val + thick_border[-1] + space for val in thick_border]
        for cell in [cell for row in sheet for cell in row]:
            if tment.t_stagger and split < cell.column <= split + space:
                cell.fill = openpyxl.styles.PatternFill(None)
            elif cell.column in thick_border or cell.column + team_width in thick_border:
                cell.border = styles.Border(right=thick if cell.column in thick_border else thin)
        for i in range(2*tment.t_pairs):
            start_col = 2 + i*team_width + (space + 1)*staggered[i // 2]
            sheet.merge_cells(start_row=1, start_column=start_col,
                              end_row=1, end_column=start_col + team_width - 1)

def export_team_views(tment, workbook, time_fmt, team_info, event_names, rooms):
    """Adds event-sorted and time-sorted team-focused views to the output workbook."""
    #writing data
    ws_chron = workbook.create_sheet("Team View (Chronological)")
    ws_event = workbook.create_sheet("Team View (Event)")
    team_header = ['Team Number'] + (['Division'] if tment.divisions else []) + ['Team Name']
    ws_chron.append(team_header + ['Event ' + str(i + 1) for i in range(len(event_names))])
    ws_event.append(team_header + event_names[2:])

    for team in sorted(tment.teams, key=lambda t: t.num):
        ws_chron.append([team.num] + team_info
                        + ['{} at {}, {}'.format(event_names[cat], time.strftime(time_fmt),
                                                 rooms[min(5, cat)][loc])
                           for (time, duration, cat, loc) in team.events])
        ws_event.append([team.num] + team_info
                        + ['{}, {}'.format(time.strftime(time_fmt), rooms[min(5, cat)][loc])
                           for (time, length, cat, loc)
                           in sorted(team.events, key=lambda x: x[2])[2:]])

    #formatting - borders, cell merges, striped shading, etc
    col_size = 1 + max(len(team.name) for team in tment.teams)
    for sheet in (ws_chron, ws_event):
        basic_sheet_format(sheet)
        for col in sheet.columns:
            if col[0].column == 2 + tment.divisions:
                sheet.column_dimensions[get_column_letter(col[0].column)].width = col_size

def basic_sheet_format(sheet, start=0):
    """bolds the top row, stripes the rows, and sets column widths"""
    for col in sheet.columns:
        col[0].font = openpyxl.styles.Font(bold=True)
        length = 2 + max((len(str(cell.value)) for cell in col[start:]
                          if cell.value and str(cell.value)[0] != '='), default=0)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = length
    row_counter = 0
    for row in sheet.rows:
        row[0].alignment = openpyxl.styles.Alignment(horizontal='right')
        row_counter = (row[0].value != '')*(row_counter + 1)
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            if cell.row > 1 and row_counter % 2:
                cell.fill = openpyxl.styles.PatternFill('solid', fgColor='DDDDDD')

def sheet_borders(sheet, borders=()):
    """Formats worksheet to add specified borders at a given period and offset."""
    for row in sheet.rows:
        for cell in row:
            for border, period, offset, start_row in borders[::-1]:
                if (cell.column - offset) % period == 0 and cell.row > start_row:
                    cell.border = border

def generate_schedule():
    """Top-most level function; gets a file, reads and schedules for it, then exports the result."""
    try:
        if len(sys.argv) == 1:
            root = tkinter.Tk()
            root.withdraw()
            fpath = filedialog.askopenfilename(initialdir=os.path.dirname(os.path.abspath(__file__)),
                                               filetypes=[("Excel files", "*.xls *.xlsm *.xlsx")])
            root.destroy()
        elif len(sys.argv) == 2:
            fpath = sys.argv[1]
        else:
            raise SystemExit("{} only accepts either zero or one files as input".format(sys.argv[0]))

        logic_params, tournament_name, io_params = read_data(fpath)
        tment = Tournament(*logic_params)
        tment.schedule()

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning)
            workbook = openpyxl.load_workbook(fpath)
        export(tment, workbook, *io_params)
        outfpath = os.path.join(os.path.dirname(fpath),
                                tournament_name.lower().replace(' ', '_') + '_schedule{}.xlsx')
        saved, attempts = False, 0
        while not saved:
            try:
                final_fout = outfpath.format(' ({})'.format(attempts) if attempts else '')
                workbook.save(final_fout)
                saved = True
            except PermissionError:
                attempts += 1
        print('Schedule saved: {}'.format(final_fout))

    except (Exception, SystemExit) as excep:
        raise excep
        #print(excep)

    #this is expected to run in console windows which close very quickly on windows
    if sys.platform in ["win32"]:
        os.system("pause")

if __name__ == "__main__":
    print("Python environment initialized")
    generate_schedule()
